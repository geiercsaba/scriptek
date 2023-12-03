from openpyxl import load_workbook, Workbook
import os
from config import logger, WORKDIR_PATH


def ha_to_m2(num):
    if '.' in num:
        m2 = num.split('.')
        return int(m2[0]) * 10000 + int(m2[1])
    return int(num)


def load_excel(file_name):
    array = []
    try:
        worksheet = load_workbook(filename=file_name).active
        for row in worksheet.iter_rows():
            if row[0].value != "Település":
                array.append([row[0].value, row[1].value])
    except FileNotFoundError:
        logger.warning(f"Nem található ilyen excel táblázat")
        raise Exception(f"Nincs ilyen file {file_name}")
    except Exception as e:
        logger.warning(f"Nem található megfelelő excel táblázat")
        raise Exception("Hiba történt az excel file beolvasása során", e)
    logger.info(f"A bemeneti Excel táblézat adatai sikeresen betöltve. Összesen: {len(array)}db hasznos sor.")
    return array


def write_to_excel(file_name, data):
    try:
        if os.path.exists(file_name):
            workbook = load_workbook(filename=file_name)
            worksheet = workbook.active
            logger.debug(f"A kimeneti Excel file már létezik, sikeresen betöltve!")
        else:
            workbook = Workbook()
            worksheet = workbook.active

            columns = ['Település', 'Hrsz', 'Művelés', 'Minőségi osztály', 'Térmérték',
                       'Aranykorona (Kat.t.jöv)', 'Érkezési idő', 'Tulajdoni hányad', 'Jogcím', 'Jogállás',
                       'Név', 'Szül.név', 'Szül.dátum', 'Anyja neve', 'Irányítószám',
                       'Helység', 'Cím', 'Teher']

            worksheet.append(columns)
            logger.debug(f"A kimeneti Excel file még nem létezik, sikeresen létrehozva, fejléc sikeresen hozzáadva!")

        count = 0
        for row in data:
            for owner in row.tulajdonosok:
                if row.muveles:
                    row_data = [row.telepules, row.hrsz, row.muveles, row.minosegi_osztaly, ha_to_m2(row.termertek),
                                row.aranykorona, owner.erkezesi_ido, owner.tulajdoni_hanyad, owner.jogcim, owner.jogallas, owner.nev, owner.szul_nev,
                                owner.szul_date, owner.anyja_neve, owner.iranyitoszam, owner.helyseg, owner.cim, row.teher]
                    worksheet.append(row_data)
                count += 1

        workbook.save(filename=file_name)
        logger.info(f"A rekordok sikeresen feltöltve az Excel táblába. Összesen: {count} db sor hozzáadva a táblázathoz")
    except Exception:
        logger.warning(f"Valamilyen váratlan hiba lépett fel az Excel file elmentése során")


def folder_to_excel(mappa):
    import os
    import glob
    from modules.pdf_data_class import Property

    shame = os.path.join(mappa, "*.pdf")
    files = glob.glob(shame)
    tullapok = []
    for file in files:
        try:
            name1 = file.split("/")[-1].split(".pdf")[0]
            name = name1.split("_")
            # Tullap_{város}_{hrsz1}_{hrsz2}
            city = name[1]
            hrsz = name[2] + "/" + name[3]
            tullapok.append(Property(city, hrsz, file))
        except IndexError:
            logger.warning(f"A {name1} file  nem illik bele a névsablonba vagy más hiba keletkezett. Manuálisan kell felvinni.")

    excel_save_path = WORKDIR_PATH + "/kimeneti_excel.xlsx"
    write_to_excel(excel_save_path, tullapok)

